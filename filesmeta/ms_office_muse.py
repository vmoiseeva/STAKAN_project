import os
import re
import zipfile
from pptx import Presentation
from openpyxl import load_workbook


class MSOfficeMuse:
    ex = [".docx", ".xlsx", ".pptx"]

    def get_meta_inf(self, path):
        metadata = {}
        ext = os.path.splitext(path)[1]

        if ext == ".docx":
            try:
                with zipfile.ZipFile(path, "r") as archive:
                    if "docProps/app.xml" in archive.namelist():
                        ms_data = archive.read("docProps/app.xml")
                        app_xml = ms_data.decode("utf-8")

                        regex = r"<(Pages)>(\d+)</(Pages)>"
                        matches = re.findall(regex, app_xml, re.MULTILINE)
                        match = matches[0] if matches[0:] else [0, 0]
                        num_pages = match[1]
                        metadata["Pages"] = num_pages
                    else:
                        print(f"File '{path}' does not contain 'docProps/app.xml'")
            except Exception as e:
                print(f"Error reading {ext} file '{path}': {e}")
        elif ext == ".pptx":
            prs = Presentation(path)
            num_slides = len(prs.slides)
            metadata["Slides"] = num_slides
        elif ext == ".xlsx":
            wb = load_workbook(path, read_only=True)
            metadata["Sheets"] = len(wb.sheetnames)

        return metadata